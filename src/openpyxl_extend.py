# encoding: utf-8

"""
扩展openpyxl,以便支持表操作
"""

from types import MethodType
from openpyxl import load_workbook


def get_all_tables(workbook):
    """ 获得所有表对象 """

    tables = []
    for sheet in workbook.worksheets:
        for table in sheet._tables:
            tables.append(table)

    return tables


def get_cell(table, row, col):
    """ 获得table中位置为(row,col)的cell """

    return table.ref_range[row][col]


def get_row(table, row):
    """ 获得第row行 """

    return table.ref_range[row]


def get_row_count(table):
    """ 获得行数 """

    ref_range = table.ref_range
    return len(ref_range)


def get_col_count(table):
    """ 获得列数 """

    return len(table.tableColumns)


def get_table_by_name(worksheet, name):
    """ 获得名字为name的表对象 """

    # 遍历sheet中的每一个表,没有找到则抛出错误
    for table in worksheet._tables:
        if table.name == name:
            return table

    return None


def load_workbook_ex(filename):
    """
    自动workbook的加载函数,从而支持加载table之类的接口
    """
    workbook = load_workbook(filename, data_only=True)

    # Workbook的函数扩展
    workbook.get_all_tables = MethodType(get_all_tables, workbook)

    # WorkSheet的函数扩展
    for sheet in workbook.worksheets:
        sheet.get_table_by_name = MethodType(get_table_by_name, sheet)

        # Table的函数扩展
        for table in sheet._tables:
            table.worksheet = sheet
            table.ref_range = sheet[table.ref]
            table.get_cell = MethodType(get_cell, table)
            table.get_row = MethodType(get_row, table)
            table.get_row_count = MethodType(get_row_count, table)
            table.get_col_count = MethodType(get_col_count, table)

    return workbook
